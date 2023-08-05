

from flask import Blueprint, jsonify, request, current_app
from app.models import Employee, Company
from app import db, create_app
from app.schemas import EmployeeSchema, CompanySchema
from openpyxl import load_workbook
import os
import random
from sqlalchemy.exc import SQLAlchemyError

bp = Blueprint('database', __name__)

@bp.route('/read-excel', methods=['GET'])
def read_excel():
    excel_filename = 'BrainerHubAssignmentSheet.xlsx'
    excel_path = os.path.join(current_app.config['UPLOAD_FOLDER'], excel_filename)

    wb = load_workbook(excel_path)
    ws = wb.active

    data = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        data.append({
            'first_name': row[1],
            'last_name': row[2],
            'phone_number': row[3]
        })

    return jsonify(data), 200

@bp.route('/insert-data', methods=['POST'])
def insert_data():
    try:
        excel_filename = 'BrainerHubAssignmentSheet.xlsx'
        excel_path = os.path.join(create_app().config['UPLOAD_FOLDER'], excel_filename)

        wb = load_workbook(excel_path)
        ws = wb.active

        employees = []
        companies = []

        for row in ws.iter_rows(min_row=2, values_only=True):
            employee_data = {
                'first_name': row[1],
                'last_name': row[2],
                'phone_number': row[3]
            }
            employee = EmployeeSchema().load(employee_data)
            employees.append(employee)

            company_data = {
                'name': 'Company ' + str(row[0]),
                'salary': random.uniform(1000, 100000),
                'manager_id': random.randint(1, 10),
                'department_id': random.randint(1, 5)
            }
            company = CompanySchema().load(company_data)
            companies.append(company)

        with db.session.begin():
            db.session.bulk_insert_mappings(Employee, employees)
            db.session.bulk_insert_mappings(Company, companies)

        return jsonify({'message': 'Data inserted successfully'}), 200
    except SQLAlchemyError as e:
        db.session.rollback()
        error_msg = str(e.__dict__.get('orig'))
        return jsonify({'error': error_msg}), 500
    except Exception as e:
        return jsonify({'error': str(e)}), 500
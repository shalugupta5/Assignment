from flask import Blueprint, request, jsonify
from flask import current_app
from app import db
import openpyxl
import os
import random
from app.models import Employee, Company
from openpyxl import Workbook
import tempfile

bp = Blueprint('excel', __name__)
def serialize_row(row):
    return {
        'id': row[0].value,
        'first_name': row[1].value,
        'last_name': row[2].value,
        'phone_number': row[3].value,
        'company_name': row[4].value,
        'salary': row[5].value,
        'manager_id': row[6].value,
        'department_id': row[7].value

    }


@bp.route('/generate-excel', methods=['GET'])
def generate_excel():
    wb = openpyxl.Workbook()
    
    ws_employee = wb.active
    ws_employee.title = 'Employee Data'
    
    employee_headers = ['id', 'first_name', 'last_name', 'phone_number', 'company_name', 'salary', 'manager_id', 'department_id']
    for col_num, header in enumerate(employee_headers, 1):
        ws_employee.cell(row=1, column=col_num, value=header)
    
    ws_company = wb.create_sheet(title='Company Data')
    
    company_headers = ['name', 'salary', 'manager_id', 'department_id']
    for col_num, header in enumerate(company_headers, 1):
        ws_company.cell(row=1, column=col_num, value=header)
    
    for row_num in range(2, 102):
        first_name = f'Employee{row_num - 1} First'
        last_name = f'Employee{row_num - 1} Last'
        phone_number = f'555-123-{row_num:03d}'
        company_name = f'Company {row_num - 1}'
        salary = round(random.uniform(30000, 100000), 2)
        
        ws_employee.cell(row=row_num, column=1, value=row_num - 1)
        ws_employee.cell(row=row_num, column=2, value=first_name)
        ws_employee.cell(row=row_num, column=3, value=last_name)
        ws_employee.cell(row=row_num, column=4, value=phone_number)
        ws_employee.cell(row=row_num, column=5, value=company_name)
        ws_employee.cell(row=row_num, column=6, value=salary)
        ws_employee.cell(row=row_num, column=7, value=random.randint(1, 10))
        ws_employee.cell(row=row_num, column=8, value=random.randint(1, 5))
        
        ws_company.cell(row=row_num, column=1, value=company_name)
        ws_company.cell(row=row_num, column=2, value=salary)
        ws_company.cell(row=row_num, column=3, value=random.randint(1, 10))
        ws_company.cell(row=row_num, column=4, value=random.randint(1, 5))
    
    excel_filename = 'BrainerHubAssignmentSheet.xlsx'
    excel_path = os.path.join(current_app.config['UPLOAD_FOLDER'], excel_filename)
    wb.save(excel_path)
    
    return jsonify({'message': 'Excel files generated successfully'}), 200

@bp.route('/read-excel', methods=['GET'])
def read_excel():
    excel_filename = 'BrainerHubAssignmentSheet.xlsx'
    excel_path = os.path.join(current_app.config['UPLOAD_FOLDER'], excel_filename)

    wb = openpyxl.load_workbook(excel_path)
    ws = wb['Employee Data']

    data = [serialize_row(row) for row in ws.iter_rows(min_row=2)]

    return jsonify(data), 200
